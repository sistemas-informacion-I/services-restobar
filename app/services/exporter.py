import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Restobar wine palette
ACCENT_COLOR = colors.HexColor("#ac111a")    # wine-700
ALT_ROW_COLOR = colors.HexColor("#fdf2f2")   # wine-50


def export_to_excel(report_label: str, category: str, columns: list, column_labels: dict, rows: list, total: int, applied_filters: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Styles
    title_font = Font(name="Segoe UI", size=14, bold=True, color="AC111A")
    meta_font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="AC111A", end_color="AC111A", fill_type="solid")
    data_font = Font(name="Segoe UI", size=10)

    # Title & Meta
    ws.append([report_label])
    ws.cell(row=1, column=1).font = title_font

    ws.append([f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.cell(row=2, column=1).font = meta_font

    ws.append([f"Categoría: {category} | Registros: {total}"])
    ws.cell(row=3, column=1).font = meta_font

    ws.append(["Filtros aplicados:"])
    ws.cell(row=4, column=1).font = meta_font

    current_row = 5
    for filter_desc in applied_filters:
        ws.cell(row=current_row, column=1, value=f"  • {filter_desc}").font = meta_font
        current_row += 1

    current_row += 1  # Empty spacer row

    # Headers
    headers = [column_labels.get(col, col) for col in columns]
    ws.append(headers)

    header_row_idx = current_row
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    current_row += 1

    # Rows
    for row_data in rows:
        row_values = [row_data.get(col, "") for col in columns]
        ws.append(row_values)
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = data_font
        current_row += 1

    # Auto-adjust column width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < header_row_idx:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_to_pdf(report_label: str, category: str, columns: list, column_labels: dict, rows: list, total: int, applied_filters: list) -> bytes:
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=ACCENT_COLOR,
        spaceAfter=6
    )

    meta_style = ParagraphStyle(
        'DocMeta',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#64748B"),
        spaceAfter=4
    )

    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor("#0F172A")
    )

    header_cell_style = ParagraphStyle(
        'HeaderCellText',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white
    )

    story = []

    # Title
    story.append(Paragraph(report_label, title_style))

    # Metadata
    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M")
    story.append(Paragraph(f"Generado: {gen_time}  |  Categoría: {category}  |  Total registros: {total}", meta_style))

    # Filters
    filters_str = "  ·  ".join(applied_filters)
    story.append(Paragraph(f"<b>Filtros aplicados:</b> {filters_str}", meta_style))
    story.append(Spacer(1, 12))

    # Table Data
    table_data = []

    header_row = [Paragraph(column_labels.get(col, col), header_cell_style) for col in columns]
    if not header_row:
        header_row = [Paragraph("Sin columnas", header_cell_style)]
    table_data.append(header_row)

    for idx, row in enumerate(rows):
        row_cells = []
        for col in columns:
            val = row.get(col, "")
            row_cells.append(Paragraph(str(val), cell_style))
        table_data.append(row_cells)

    if not columns:
        table_data.append([Paragraph("No se seleccionaron columnas para mostrar", cell_style)])

    col_widths = None
    t = Table(table_data, colWidths=col_widths, repeatRows=1)

    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), ACCENT_COLOR),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
    ])

    for i in range(1, len(table_data)):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), ALT_ROW_COLOR)

    t.setStyle(t_style)
    story.append(t)

    doc.build(story)
    return output.getvalue()


def export_to_html(report_label: str, category: str, columns: list, column_labels: dict, rows: list, total: int, applied_filters: list) -> str:
    def esc(s):
        if s is None:
            return ""
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

    head_parts = []
    for col in columns:
        head_parts.append(f"<th>{esc(column_labels.get(col, col))}</th>")

    body_parts = []
    for row in rows:
        row_str = "<tr>"
        for col in columns:
            val = row.get(col, "")
            row_str += f"<td>{esc(val)}</td>"
        row_str += "</tr>"
        body_parts.append(row_str)

    filters_list = "".join([f"<li>{esc(f)}</li>" for f in applied_filters])
    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M")

    html = f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<title>{esc(report_label)}</title>
<style>
  body{{font-family:'Segoe UI',Arial,sans-serif;padding:28px;color:#0F172A;background:#fff;}}
  h1{{color:#ac111a;margin:0 0 4px;font-size:20px;}}
  .meta{{color:#64748B;font-size:12px;margin:2px 0;}}
  .filters{{background:#fdf2f2;border:1px solid #fbcad0;border-radius:8px;
           padding:10px 14px;margin:14px 0;font-size:12px;}}
  .filters b{{color:#820a12;}}
  .filters ul{{margin:6px 0 0;padding-left:18px;}}
  table{{border-collapse:collapse;width:100%;margin-top:10px;}}
  th{{background:#ac111a;color:#fff;padding:8px 10px;text-align:left;font-size:12px;}}
  td{{border-bottom:1px solid #E2E8F0;padding:6px 10px;font-size:12px;}}
  tr:nth-child(even) td{{background:#fdf2f2;}}
  @media print{{button{{display:none;}}}}
</style></head><body>
<h1>{esc(report_label)}</h1>
<div class="meta">Generado: {gen_time}</div>
<div class="meta">Tipo: {esc(category)} &nbsp;|&nbsp; Total registros: {total}</div>
<div class="filters"><b>Filtros aplicados</b><ul>{filters_list}</ul></div>
<table><thead><tr>{"".join(head_parts)}</tr></thead><tbody>{"".join(body_parts)}</tbody></table>
</body></html>"""
    return html
